#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""DeepCable-Grandmaster - Generate formatted Excel audit report for 得力电缆.xlsx"""

import re
import math
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# ============================================================
# RAW DATA
# ============================================================
raw_entries = [
    ("YJV-10KV 3x300", 32.48),
    ("YJV-4X240+1X120", 27.6),
    ("BBTRZ-5X6", 886.6),
    ("NHYJV-0.6/1KV- 5X6", 314.97),
    ("NHYJV-0.6/1KV- 5X4", 202.2),
    ("NHYJV-0.6/1KV-3x4", 726.9),
    ("YJV-0.6/1KV 5X10", 93.56),
    ("YJV-0.6/1KV 5X16", 1167.16),
    ("YJV-0.6/1KV 4X25+1X16", 48.8),
    ("BV2.5", 24086.04),
    ("BVR2.5", 11894.83),
    ("NH-BV2.5", 416.79),
    ("NH-BVR2.5", 163.11),
    ("NH-BV4", 5.44),
    ("NH-BVR4", 2.72),
    ("NH-RVS-2x2.5", 33.26),
    ("WDZN-RYSP-2X1.5", 24.84),
    ("WDZN-RYJS-2x2.5", 6804.2),
    ("BBTRZ-5X6", 865.3),
    ("NHYJV-0.6/1KV- 5X6", 865.3),
    ("NHYJV-0.6/1KV- 5X4", 271.2),
    ("NHYJV-0.6/1KV-3x4", 1068.4),
    ("YJV-0.6/1KV 5X4", 17.7),
    ("YJV-0.6/1KV 5X6", 32.74),
    ("YJV-0.6/1KV 5X16", 2950.58),
    ("YJV-0.6/1KV 4X25+1X16", 612.46),
    ("YJV-0.6/1KV 4X35+1X16", 1006.16),
    ("NH-YJV-0.6/1KV 3X25", 93.45),
    ("NH-YJV-0.6/1KV 3X25+1X16", 93.45),
    ("NH-YJV-0.6/1KV 3X95", 42.68),
    ("NH-YJV-0.6/1KV 3X95+1X50", 42.68),
    ("BV1.5", 18.62),
    ("BV2.5", 35123.68),
    ("BVR2.5", 17296.2),
    ("NH-BV2.5", 527.5),
    ("NH-BVR2.5", 221.36),
    ("WDZN-RYSP-2X1.5", 80.63),
    ("WDZN-RYJS-2x2.5", 9483.89),
    ("BBTRZ-5X6", 876.61),
    ("NHYJV-0.6/1KV- 5X6", 876.61),
    ("NHYJV-0.6/1KV- 5X4", 144.73),
    ("NHYJV-0.6/1KV-3x4", 1114.25),
    ("YJV-0.6/1KV 5X10", 221.04),
    ("YJV-0.6/1KV 5X16", 1840.61),
    ("YJV-0.6/1KV 4X25+1X16", 1215.22),
    ("YJV-0.6/1KV 4X35+1X16", 691.47),
    ("BV2.5", 43450.52),
    ("BVR2.5", 20844.83),
    ("NH-BV2.5", 215.32),
    ("NH-BVR2.5", 68.55),
    ("WDZN-RYSP-2X1.5", 138.13),
    ("WDZN-RYJS-2x2.5", 9028.81),
    ("BV2.5", 293.42),
    ("BVR2.5", 126.81),
    ("BV2.5", 152.8),
    ("BVR2.5", 76.4),
    ("NH-BV2.5", 757.51),
    ("NH-BVR2.5", 273.19),
    ("NH-RVS-2x2.5", 115.14),
    ("BBTRZ-3X4", 16.41),
    ("NHYJV-0.6/1KV- 3X6", 88.99),
    ("WDZBN-YJY-0.6/1KV 3X4", 0.73),
    ("YJV-0.6/1KV 3X6", 1162.01),
    ("YJV-0.6/1KV 5X4", 10.31),
    ("YJV-0.6/1KV 5X6", 261.33),
    ("YJV-0.6/1KV 5X10", 178.98),
    ("YJV-0.6/1KV 5X16", 59.2),
    ("YJV-0.6/1KV 4X25+1X16", 125.86),
    ("YJV-0.6/1KV 4X70+1X35", 114.66),
    ("BV1.5", 404.64),
    ("BV2.5", 519.6),
    ("BVR2.5", 221.21),
    ("NHBV2.5", 32.98),
    ("NHBVR2.5", 16.49),
    ("WDZN-RYJS-2x2.5", 848.09),
    ("BBTRZ-3X4", 42.83),
    ("NHYJV-0.6/1KV- 3X6", 187.31),
    ("WDZBN-YJY-0.6/1KV 3X4", 3.29),
    ("YJV-0.6/1KV 5X10", 298.24),
    ("YJV-0.6/1KV 4X25+1X16", 186.25),
    ("YJV-0.6/1KV 4X35+1X16", 175.99),
    ("BV2.5", 438.46),
    ("BVR2.5", 210.28),
    ("NH-BV2.5", 10.96),
    ("NH-BVR2.5", 5.48),
    ("WDZN-RYJS-2x2.5", 974.53),
    ("NH-RVS-2x2.5", 11.92),
    ("NHYJV-0.6/1KV 3X6", 26.16),
    ("WDZBN-YJY-0.6/1KV 3X4", 43.22),
    ("YJV-0.6/1KV 5X4", 32.03),
    ("BV2.5", 3622.3),
    ("BVR2.5", 1786.78),
    ("NH-BV2.5", 28.74),
    ("NH-BVR2.5", 12.24),
    ("WDZN-RYJS-2x2.5", 1320.2),
    ("YJV-0.6/1KV 3X4", 1276.25),
    ("YJV-0.6/1KV 3X6", 1237.94),
    ("NHYJV-0.6/1KV 5X16", 2678.33),
    ("NHYJV-0.6/1KV 4X95+1X50", 165.48),
    ("NHYJV-0.6/1KV 4X120+1X70", 775.81),
    ("YJV-0.6/1KV 5X16", 105.7),
    ("YJV-0.6/1KV 4X35+1X16", 119.68),
    ("YJV-0.6/1KV 4X70+1X35", 1637.37),
    ("YJV-0.6/1KV 4X120+1X70", 67.44),
    ("YJV-0.6/1KV 4X150+1X70", 194.54),
    ("YJV22-0.6/1KV 4X25+1X16", 633.89),
    ("YJV22-0.6/1KV 4X50+1X25", 633.89),
    ("YJV22-0.6/1KV 4X95+1X50", 304.51),
    ("YJV22-0.6/1KV 4X120+1X70", 609.01),
    ("YJV22-0.6/1KV 4X150+1X70", 473.16),
    ("YJV22-0.6/1KV 4X185+1X95", 304.51),
    ("YJV22-0.6/1KV 5X16", 329.13),
    ("WDZN-RYJSP 2*2.5", 384.23),
    ("WDZN-RYJSP 2*2.5", 711.3),
    ("WDZN-RYSP 2*1.5", 446.63),
    ("WDZN-RYSP 2*1.5", 452.04),
    ("WDZN-KYJYP2*1.5", 191.43),
    ("WDZN-KYJY2*2.5", 192.99),
    ("WDZN-KYJY10*1.5", 192.99),
    ("WDZN-RYJS 2*1.5", 7108.94),
    ("WDZN-RYS 2*1.5", 1702.33),
    ("WDZN-BYJ 2.5", 423.38),
    ("WDZN-YJY 2*6", 192.99),
    ("WDZN-RYJSP 2*2.5", 667.91),
    ("WDZN-RYJSP 2*2.5", 664.51),
    ("WDZN-RYSP 2*1.5", 281.07),
    ("WDZN-RYSP 2*1.5", 497.08),
    ("WDZN-RYJS 2*1.5", 11681.21),
    ("WDZN-RYS 2*1.5", 2974.98),
    ("WDZN-BYJ 2.5", 557.78),
    ("WDZN-KYJY8*1.5", 965.97),
    ("WDZN-KYJY7*1.5", 952.55),
    ("WDZN-KYJY2*2.5", 540.39),
    ("WDZN-RYJSP 2*2.5", 863.48),
    ("WDZN-RYJSP 2*2.5", 748.55),
    ("WDZN-RYSP 2*1.5", 535.03),
    ("WDZN-RYSP 2*1.5", 549.4),
    ("WDZN-RYS 2*1.5", 2906.56),
    ("WDZN-RYJS 2*1.5", 9245.13),
    ("WDZN-BYJ 2.5", 1010.69),
    ("WDZN-KYJY2*2.5", 30.36),
    ("WDZN-KYJY12*1.5", 139.58),
    ("WDZN-KYJY2*1.5", 139.58),
    ("WDZN-KYJY5*1.5", 135.62),
    ("WDZN-YJY 2*6", 135.62),
    ("WDZN-YJY 2*2.5", 168.38),
    ("WDZN-KYJY 3*1.5", 168.38),
    ("WDZN-BYJ 2.5", 48.84),
    ("WDZN-RYJS 2*1.5", 131.41),
    ("WDZN-RYJS 2*1.5", 160.57),
    ("WDZN-KYJY 3*1.5", 154.26),
    ("WDZN-RYSP-2x1.5", 19.95),
    ("WDZN-RYSP-2x1.5", 2.29),
    ("WDZN-RYSP-2x1.5", 2.72),
    ("WDZN-RYJS-2x1.5", 168.1),
    ("WDZN-RYSP 2*1.5", 25.36),
    ("WDZN-RYSP 2*1.5", 4.52),
    ("WDZN-RYJS 2*1.5", 117.35),
    ("WDZN-RYSP-2x1.5", 147.06),
    ("WDZN-RYSP-2x1.5", 145.78),
    ("WDZN-RYSP-2x1.5", 190.03),
    ("WDZN-RYJS 2*1.5", 338.88),
    ("WDZN-KYJYP2*1.5", 97.48),
    ("WDZN-KYJYP3*1.5", 434.1),
    ("WDZN-KYJY2*2.5", 625.71),
    ("WDZN-KYJY7*1.5", 293.65),
    ("WDZN-KYJY8*1.5", 293.93),
    ("WDZN-KYJY10*1.5", 96.23),
    ("WDZN-KYJY16*1.5", 194.24),
    ("WDZN-YJY 2*2.5", 217.02),
    ("WDZN-YJY 2*6", 290.27),
    ("WDZN-RYSP 2*1.5", 846.055),
    ("WDZN-RYSP 2*2.5", 95.115),
    ("WDZN-RYJSP 2*2.5", 258.9),
    ("BV2.5", 5020.47),
    ("BVR2.5", 2065.2),
    ("BV2.5", 5246.69),
    ("BVR2.5", 2169.34),
    ("BV2.5", 4629.51),
    ("BVR2.5", 1916.08),
    ("BV2.5", 2460.15),
    ("BVR2.5", 947.25),
    ("BV2.5", 5552.66),
    ("BVR2.5", 1673.14),
    ("BV4", 470.84),
    ("BVR4", 235.42),
    ("BV2.5", 9290.17),
    ("BVR2.5", 4068.22),
    ("YJV-0.6/1KV 4X25+1X16", 220.87),
    ("BV2.5", 248.75),
    ("BVR2.5", 103.19),
    ("BV4", 36.64),
    ("BVR4", 18.32),
    ("BV2.5", 320.01),
    ("BVR2.5", 134.59),
    ("BV2.5", 148.39),
    ("BVR2.5", 56.93),
]

# ============================================================
# PE TABLE
# ============================================================
PE_TABLE = {
    2.5: 1.5, 4: 2.5, 6: 4, 10: 6, 16: 10, 25: 16, 35: 16,
    50: 25, 70: 35, 95: 50, 120: 70, 150: 70, 185: 95, 240: 120,
    300: 150, 400: 185
}

ALL_CORES = ['YJV22', 'YJV', 'YJY', 'VV', 'KYJYP', 'KYJY', 'KVV', 'KVVP',
             'RYJSP', 'RYJS', 'RVS', 'RVSP', 'BYJR', 'BYJ', 'BVR', 'BV', 'BBTRZ']

POWER_CORES = {'YJV22', 'YJV', 'YJY', 'VV', 'BBTRZ'}
CONTROL_CORES = {'KYJYP', 'KYJY', 'KVV', 'KVVP'}
TWISTED_CORES = {'RYJSP', 'RYJS', 'RVS', 'RVSP'}
WIRE_CORES = {'BYJR', 'BYJ', 'BVR', 'BV'}

VOLTAGE_MAP = {
    'POWER': '0.6/1kV',
    'CONTROL': '450/750V',
    'TWISTED': '300/300V',
    'BYJ': '450/750V',
    'BVR': '450/750V',
    'BV_GE_1.5': '450/750V',
    'BV_LE_1.0': '300/500V',
}


def check_n_pe(std_model):
    warnings = []
    m = re.search(r'(?:kV|V)-(.+)$', std_model)
    if not m:
        return warnings
    sec = m.group(1)
    p = re.search(r'(\d+)×(\d+\.?\d*)\+(\d+)×(\d+\.?\d*)', sec)
    if p:
        ps = float(p.group(2))
        ns = float(p.group(4))
        for k in sorted(PE_TABLE.keys()):
            if ps <= k:
                if ns < PE_TABLE[k]:
                    warnings.append(
                        f'[PE报警] 相线{ps}mm²→PE需≥{PE_TABLE[k]}mm²，实配{ns}mm²'
                    )
                break
    return warnings


def extract_components(m):
    comps = {'prefix': '', 'core': '', 'voltage': '', 'section': '',
             'core_count': 0, 'pe_section': '', 'b1': '', 'is_mv': False}

    # RYS→RYJS, RYSP→RYJSP
    m = re.sub(r'\bRYS\b', 'RYJS', m)
    m = m.replace('RYSP', 'RYJSP')
    # NH→N
    m = re.sub(r'\bNH-', 'N-', m)
    m = re.sub(r'\bNH(?=[A-Z])', 'N', m)
    m = re.sub(r'\bNHBV', 'N-BV', m)
    m = re.sub(r'\bNHBVR', 'N-BVR', m)
    m = re.sub(r'\bNHRVS', 'N-RVS', m)
    m = re.sub(r'\bNYJV22\b', 'N-YJV22', m)
    m = re.sub(r'\bNYJV\b', 'N-YJV', m)
    # * and x to ×
    m = re.sub(r'\*', '×', m)
    m = re.sub(r'(?<=\d)[xX](?=\d)', '×', m)
    # Voltage case
    m = re.sub(r'(\d)[Kk][Vv]', lambda x: x.group(1).lower() + 'kV', m)
    # Clean spaces
    m = re.sub(r'-\s+', '-', m)
    m = re.sub(r'\s+', '', m)
    m = re.sub(r'-+', '-', m)
    m = m.strip('-')

    # B1
    if '-B1-' in m:
        comps['b1'] = 'B1'
        m = m.replace('-B1-', '')

    # Detect MV
    if re.search(r'\b(\d+)kV\b', m):
        kv = int(re.search(r'\b(\d+)kV\b', m).group(1))
        if kv >= 6:
            comps['is_mv'] = True

    # Extract prefix
    prefixes = sorted(['WDZBN', 'WDZCN', 'WDZN', 'WDZB', 'WDZC', 'WDZ', 'WDUZC', 'WDUZ', 'N'],
                      key=len, reverse=True)
    for p in prefixes:
        if m.startswith(p):
            comps['prefix'] = p
            m = m[len(p):]
            if m.startswith('-'):
                m = m[1:]
            break

    # Extract voltage
    vm = re.search(r'(DC\d+V|\d+\.?\d*/\d+\.?\d*kV|\d+kV|\d+/\d+V)', m)
    if vm:
        comps['voltage'] = vm.group(1)
        m = m[:vm.start()] + m[vm.end():]
        if m.startswith('-'):
            m = m[1:]

    # Extract core type
    core_found = ''
    for c in sorted(ALL_CORES, key=len, reverse=True):
        if m.startswith(c):
            core_found = c
            m = m[len(c):]
            break
    if not core_found:
        for c in sorted(ALL_CORES, key=len, reverse=True):
            if c in m:
                core_found = c
                m = m.replace(c, '', 1)
                break

    comps['core'] = core_found
    if m.startswith('-'):
        m = m[1:]

    # Section info
    sec_match = re.search(r'(\d+)×(\d+\.?\d*)(?:\+(\d+)×(\d+\.?\d*))?', m)
    if sec_match:
        comps['core_count'] = int(sec_match.group(1))
        comps['section'] = sec_match.group(2)
        if sec_match.group(3) and sec_match.group(4):
            comps['pe_section'] = f'+{sec_match.group(3)}×{sec_match.group(4)}'
    else:
        bare = re.search(r'^(\d+\.?\d*)$', m)
        if bare:
            comps['section'] = bare.group(1)

    return comps


def rebuild(comps):
    prefix = comps['prefix']
    core = comps['core']
    cc = comps['core_count']
    sec = comps['section']
    pe = comps['pe_section']
    voltage = comps['voltage']
    b1 = comps['b1']
    is_mv = comps['is_mv']

    if not voltage and not is_mv:
        if core in POWER_CORES:
            voltage = '0.6/1kV'
        elif core in CONTROL_CORES:
            voltage = '450/750V'
        elif core in TWISTED_CORES:
            voltage = '300/300V'
        elif core == 'BYJ' or core == 'BYJR':
            voltage = '450/750V'
        elif core == 'BVR':
            voltage = '450/750V'
        elif core == 'BV':
            if sec:
                s = float(sec)
                voltage = '300/500V' if s <= 1.0 else '450/750V'
            else:
                voltage = '450/750V'
    elif is_mv:
        pass

    sec_str = sec
    if cc > 0:
        sec_str = f'{cc}×{sec}'
        if pe:
            sec_str += pe
    elif sec:
        sec_str = f'1×{sec}'
    else:
        sec_str = '?'

    parts = []
    if prefix:
        parts.append(prefix)
    parts.append(core)
    if b1:
        parts.append(f'B1({b1})')
    if voltage:
        parts.append(voltage)
    parts.append(sec_str)

    return '-'.join(parts)


def build_fingerprint(std):
    parts = std.split('-')
    prefix = ''
    core = ''
    b1 = ''
    voltage = ''
    section = ''

    prefixes = sorted(['WDZBN', 'WDZCN', 'WDZN', 'WDZB', 'WDZC', 'WDZ', 'WDUZC', 'WDUZ', 'N'],
                      key=len, reverse=True)
    for p in prefixes:
        if std.startswith(p):
            prefix = p
            std = std[len(p):]
            if std.startswith('-'):
                std = std[1:]
            break

    if '-B1-' in std:
        b1 = 'B1'
        std = std.replace('-B1-', '')

    vm = re.search(r'(DC\d+V|\d+\.?\d*/\d+\.?\d*kV|\d+kV)', std)
    if vm:
        voltage = vm.group(1)
        std = std[:vm.start()] + std[vm.end():]

    std = std.lstrip('-')
    for c in sorted(ALL_CORES, key=len, reverse=True):
        if std.startswith(c):
            core = c
            std = std[len(c):]
            break

    std = std.lstrip('-')
    section = std if std else '?'

    return f'{prefix}|{core}|{b1}|{voltage}|{section}'


def fmt_name(fp):
    p, c, b1, v, s = fp.split('|')
    parts = []
    if p: parts.append(p)
    parts.append(c)
    if b1: parts.append(b1)
    if v: parts.append(v)
    parts.append(s)
    return '-'.join(parts)


# ============================================================
# PROCESS ALL DATA
# ============================================================
groups = defaultdict(list)
results = []

for i, (model, qty) in enumerate(raw_entries):
    comps = extract_components(model)
    std = rebuild(comps)
    fp = build_fingerprint(std)
    warns = check_n_pe(std)
    groups[fp].append((i, qty))
    results.append((i, model, std, fp, warns))


def sort_key(item):
    fp, entries = item
    parts = fp.split('|')
    core = parts[1]
    voltage = parts[3]

    if core in POWER_CORES:
        cat = 1
        if voltage and not voltage.startswith('0.6/'):
            cat = 0
    elif core in CONTROL_CORES:
        cat = 2
    elif core == 'BBTRZ':
        cat = 3
    elif core in {'BYJ', 'BYJR'}:
        cat = 4
    elif core in {'BV', 'BVR'}:
        cat = 5
    elif core in TWISTED_CORES:
        cat = 6
    else:
        cat = 7

    sec = parts[4]
    sm = re.search(r'(\d+\.?\d*)', sec)
    sv = float(sm.group(1)) if sm else 0

    return (cat, -sv)


sorted_groups = sorted(groups.items(), key=sort_key)

# ============================================================
# CATEGORY NAMES
# ============================================================
CATEGORY_NAMES = {
    0: '中压电力电缆',
    1: '低压电力电缆',
    2: '控制电缆',
    3: '矿物绝缘电缆(BBTRZ)',
    4: '无卤布电线(BYJ/BYJR)',
    5: '普通布电线(BV/BVR)',
    6: '双绞线/屏蔽线',
    7: '其他',
}


def get_category(fp):
    parts = fp.split('|')
    core = parts[1]
    voltage = parts[3]
    if core in POWER_CORES:
        if voltage and not voltage.startswith('0.6/'):
            return 0
        return 1
    elif core in CONTROL_CORES:
        return 2
    elif core == 'BBTRZ':
        return 3
    elif core in {'BYJ', 'BYJR'}:
        return 4
    elif core in {'BV', 'BVR'}:
        return 5
    elif core in TWISTED_CORES:
        return 6
    else:
        return 7


# ============================================================
# EXCEL GENERATION
# ============================================================
wb = openpyxl.Workbook()

# Styles
header_font = Font(name='Microsoft YaHei', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
header_align = Alignment(horizontal='center', vertical='center')

cat_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
cat_font = Font(name='Microsoft YaHei', bold=True, size=11, color='2F5496')

data_font = Font(name='Microsoft YaHei', size=10)
data_align = Alignment(horizontal='center', vertical='center')
data_align_left = Alignment(horizontal='left', vertical='center')

thin_border = Border(
    left=Side(style='thin', color='B4C6E7'),
    right=Side(style='thin', color='B4C6E7'),
    top=Side(style='thin', color='B4C6E7'),
    bottom=Side(style='thin', color='B4C6E7'),
)

total_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
total_font = Font(name='Microsoft YaHei', bold=True, size=11)

pe_warn_fill = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')
pe_warn_font = Font(name='Microsoft YaHei', size=10, color='C62828')


def write_sheet_header(ws, row, headers):
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border


def write_data_row(ws, row, values, fonts=None, fills=None, aligns=None):
    for col_idx, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=col_idx, value=v)
        cell.font = fonts[col_idx - 1] if fonts else data_font
        cell.fill = fills[col_idx - 1] if fills else PatternFill()
        cell.alignment = aligns[col_idx - 1] if aligns else data_align
        cell.border = thin_border


# ========================
# Sheet 1: 标准化清单
# ========================
ws1 = wb.active
ws1.title = '标准化清单'

headers1 = ['序号', '规格型号', '单位', '数量']
write_sheet_header(ws1, 1, headers1)

row_num = 2
seq = 1
current_cat = None
grand_total = 0

for fp, entries in sorted_groups:
    cat = get_category(fp)
    if cat != current_cat:
        # Category header row
        ws1.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=4)
        cell = ws1.cell(row=row_num, column=1, value=CATEGORY_NAMES.get(cat, '其他'))
        cell.font = cat_font
        cell.fill = cat_fill
        cell.alignment = Alignment(horizontal='left', vertical='center')
        for c in range(1, 5):
            ws1.cell(row=row_num, column=c).border = thin_border
        row_num += 1
        current_cat = cat

    dn = fmt_name(fp)
    total = sum(e[1] for e in entries)
    total_rounded = math.ceil(total * 100) / 100

    write_data_row(ws1, row_num, [seq, dn, 'm', total_rounded])
    ws1.cell(row=row_num, column=2).alignment = data_align_left
    row_num += 1
    seq += 1
    grand_total += total_rounded

# Grand total row
ws1.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=3)
cell = ws1.cell(row=row_num, column=1, value='合计')
cell.font = total_font
cell.fill = total_fill
cell.alignment = Alignment(horizontal='right', vertical='center')
for c in range(1, 5):
    ws1.cell(row=row_num, column=c).border = thin_border
    ws1.cell(row=row_num, column=c).fill = total_fill
ws1.cell(row=row_num, column=4, value=round(grand_total, 2)).font = total_font

# Column widths
ws1.column_dimensions['A'].width = 8
ws1.column_dimensions['B'].width = 55
ws1.column_dimensions['C'].width = 8
ws1.column_dimensions['D'].width = 15

# ========================
# Sheet 2: 合并验算报告
# ========================
ws2 = wb.create_sheet('合并验算报告')

headers2 = ['序号', '规格型号', '单位', '数量', '合并条目明细']
write_sheet_header(ws2, 1, headers2)

row_num = 2
seq = 1
current_cat = None
grand_total2 = 0

for fp, entries in sorted_groups:
    cat = get_category(fp)
    if cat != current_cat:
        ws2.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=5)
        cell = ws2.cell(row=row_num, column=1, value=CATEGORY_NAMES.get(cat, '其他'))
        cell.font = cat_font
        cell.fill = cat_fill
        cell.alignment = Alignment(horizontal='left', vertical='center')
        for c in range(1, 6):
            ws2.cell(row=row_num, column=c).border = thin_border
        row_num += 1
        current_cat = cat

    dn = fmt_name(fp)
    total = sum(e[1] for e in entries)
    total_rounded = math.ceil(total * 100) / 100
    details = ' + '.join(f'#{idx+1}({qty}m)' for idx, qty in entries)

    write_data_row(ws2, row_num, [seq, dn, 'm', total_rounded, details],
                   aligns=[data_align, data_align_left, data_align, data_align, Alignment(horizontal='left', vertical='center', wrap_text=True)])
    row_num += 1
    seq += 1
    grand_total2 += total_rounded

# Grand total
ws2.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=3)
cell = ws2.cell(row=row_num, column=1, value='合计')
cell.font = total_font
cell.fill = total_fill
cell.alignment = Alignment(horizontal='right', vertical='center')
for c in range(1, 6):
    ws2.cell(row=row_num, column=c).border = thin_border
    ws2.cell(row=row_num, column=c).fill = total_fill
ws2.cell(row=row_num, column=4, value=round(grand_total2, 2)).font = total_font

# Summary rows
row_num += 2
info_data = [
    ('原始条目数', len(raw_entries)),
    ('合并后行数', len(groups)),
    ('合并减少', len(raw_entries) - len(groups)),
]
for label, val in info_data:
    ws2.cell(row=row_num, column=1, value=label).font = total_font
    ws2.cell(row=row_num, column=2, value=val).font = data_font
    row_num += 1

# PE warnings
pe_all = [(i+1, orig, std, w) for i, orig, std, fp, warns in results for w in warns if 'PE' in w]
if pe_all:
    row_num += 1
    ws2.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=5)
    cell = ws2.cell(row=row_num, column=1, value='PE截面报警')
    cell.font = Font(name='Microsoft YaHei', bold=True, size=11, color='C62828')
    cell.fill = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')
    row_num += 1
    for idx, orig, std, w in pe_all:
        write_data_row(ws2, row_num, [f'条目{idx}', orig, '', '', w],
                       fonts=[pe_warn_font]*5,
                       fills=[pe_warn_fill]*5,
                       aligns=[data_align, data_align_left, data_align, data_align, Alignment(horizontal='left', vertical='center')])
        row_num += 1

# Column widths
ws2.column_dimensions['A'].width = 8
ws2.column_dimensions['B'].width = 55
ws2.column_dimensions['C'].width = 8
ws2.column_dimensions['D'].width = 15
ws2.column_dimensions['E'].width = 60

# ========================
# Sheet 3: 原始数据索引
# ========================
ws3 = wb.create_sheet('原始数据索引')

headers3 = ['序号', '原始型号', '标准化型号', '数量', 'PE校验']
write_sheet_header(ws3, 1, headers3)

for i, orig, std, fp, warns in results:
    ws = '; '.join(warns) if warns else 'OK'
    write_data_row(ws3, i + 2, [i + 1, orig, std, raw_entries[i][1], ws],
                   aligns=[data_align, data_align_left, data_align_left, data_align, data_align])
    if warns:
        for c in range(1, 6):
            ws3.cell(row=i + 2, column=c).font = pe_warn_font
            ws3.cell(row=i + 2, column=c).fill = pe_warn_fill

ws3.column_dimensions['A'].width = 8
ws3.column_dimensions['B'].width = 45
ws3.column_dimensions['C'].width = 60
ws3.column_dimensions['D'].width = 15
ws3.column_dimensions['E'].width = 50

# ============================================================
# SAVE
# ============================================================
output_path = 'C:/Users/yanli/Desktop/得力电缆_审计结果.xlsx'
wb.save(output_path)
print(f'Excel文件已生成：{output_path}')
print(f'原始条目：{len(raw_entries)}条 → 合并后：{len(groups)}行（已合并{len(raw_entries) - len(groups)}条）')
print(f'合计总长度：{round(grand_total, 2)}m')
