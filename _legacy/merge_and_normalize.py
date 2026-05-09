import sys, re
import pandas as pd
sys.path.insert(0, r'C:\Users\yanli\.claude\skills\deepcable-audit')
from deepcable_normalize import normalize
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import quote_sheetname

def clean_and_normalize(spec):
    if pd.isna(spec):
        return ''
    s = str(spec).strip()
    if not s:
        return ''
    cleaned = re.sub(r'\s+', '', s)
    cleaned = re.sub(r'[xX]', '×', cleaned)
    cleaned = re.sub(r'0\.5KV', '450/750V', cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r'[Mm][Mm][2²]?$', '', cleaned)
    result = normalize(cleaned)
    return result if result else ''

path1 = r'C:\Users\yanli\Desktop\20260508-紧急采购-容桂脱水机房电气安装施工材料采购申请(1).xlsx'
path2 = r'C:\Users\yanli\Desktop\20260508-紧急采购-大门脱水机房电气安装施工材料采购申请(1).xlsx'

df1 = pd.read_excel(path1, header=None)
df2 = pd.read_excel(path2, header=None)

def trim_data(df):
    last = len(df) - 1
    while last > 3:
        row = df.iloc[last]
        if pd.notna(row[0]) or pd.notna(row[4]):
            break
        last -= 1
    return df.iloc[:last+1]

df1 = trim_data(df1)
df2 = trim_data(df2)

ronggui_data = [df1.iloc[i].tolist() for i in range(4, len(df1))]
damen_data = [df2.iloc[i].tolist() for i in range(4, len(df2))]

wb = load_workbook(path1)
ws = wb.active
ws.title = '合并采购清单'

# Unmerge ALL cells to avoid MergedCell issues
for mr in list(ws.merged_cells.ranges):
    ws.unmerge_cells(str(mr))

# Insert column F for 标准型号
ws.insert_cols(6)

# New column header
thin = Side(style='thin', color='CCCCCC')
bd = Border(left=thin, right=thin, top=thin, bottom=thin)
hdr_fill = PatternFill('solid', fgColor='1A3A6B')
hdr_font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
ca = Alignment(horizontal='center', vertical='center')
c = ws.cell(row=4, column=6, value='标准型号')
c.font = hdr_font; c.fill = hdr_fill; c.alignment = ca; c.border = bd

# Clear all old data rows
for r in range(5, ws.max_row + 1):
    for c_idx in range(1, ws.max_column + 1):
        try:
            ws.cell(row=r, column=c_idx).value = None
        except (AttributeError, KeyError):
            pass

# Write merged data
seq = 1
row_num = 5
all_data = ronggui_data + damen_data
for row in all_data:
    row[0] = seq
    for ci in range(5):
        ws.cell(row=row_num, column=ci+1, value=row[ci])
    ws.cell(row=row_num, column=6, value=clean_and_normalize(row[4]))
    for ci in range(5, len(row)):
        ws.cell(row=row_num, column=ci+2, value=row[ci])
    row_num += 1
    seq += 1

for col_letter, w in [('A',6),('B',12),('C',22),('D',22),('E',28),('F',32),('G',6),('H',8),('I',14),('J',18),('K',12)]:
    ws.column_dimensions[col_letter].width = w

output = r'C:\Users\yanli\Desktop\合并采购清单_标准化_kv.xlsx'
wb.save(output)
total = len(ronggui_data) + len(damen_data)
print(f'OK: 容桂 {len(ronggui_data)} 条 + 大门 {len(damen_data)} 条 = {total} 条')
print(f'保存至: {output}')
