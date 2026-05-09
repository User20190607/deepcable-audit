#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
process_indonesia.py
处理：报价表-印尼数据中心项目工程清单-0308 上海 -电缆汇总.xlsx

从两个位置提取电缆数据：
  Sheet: 配电系统（机房楼外）— 描述性文字中提取型号
  Sheet2: 机房区域 ZA-RVV 电缆
"""

import re
import math
import sys
import os
from collections import defaultdict

# ── 数据定义 ──────────────────────────────────────────────────────────

# 配电系统（机房楼外）- 从描述中提取
OUTDOOR_CABLES = [
    ("YJV-0.6/1kv-4x185+1x95", 1844),
    ("YJV-0.6/1kv-4x150+1x70", 1100),
    ("YJV-0.6/1kv-4x120+1x70", 4656),
    ("YJV-0.6/1kv-4x95+1x50",  2256),
    ("YJV-0.6/1kv-4x70+1x35",  576),
    ("YJV-0.6/1kv-4x25+1x16",  576),
    ("NG-A-0.6/1kv-4X70+1X35", 1120),
    ("NG-A-0.6/1kv-5X16",      796),
    ("YJV22-20kv-3x300",       3800),
    ("YJV22-20kv-3x185",       3670),
]

# 机房（Sheet2）- ZA-RVV
INDOOR_CABLES = [
    ("ZA-RVV 1x150",      36),
    ("ZA-RVV 1x16",       846),
    ("ZA-RVV 4x10+1x6",   588),
    ("ZA-RVV 4x120+1x70", 208),
    ("ZA-RVV 4x16+1x10",   72),
    ("ZA-RVV 4x185+1x95",  78),
    ("ZA-RVV 4x25+1x16",   56),
    ("ZA-RVV 4x6+1x4",    960),
]

ALL_ENTRIES = OUTDOOR_CABLES + INDOOR_CABLES

# ── 标准化处理 ────────────────────────────────────────────────────────

# 核心型号列表
ALL_CORES = ['YJV22', 'YJV', 'YJY', 'VV', 'KYJYP', 'KYJY', 'KVV', 'KVVP',
             'RYJSP', 'RYJS', 'RVS', 'RVSP', 'BYJR', 'BYJ', 'BVR', 'BV',
             'BBTRZ', 'RVV', 'RVVP']

POWER_CORES = {'YJV22', 'YJV', 'YJY', 'VV', 'BBTRZ'}
CONTROL_CORES = {'KYJYP', 'KYJY', 'KVV', 'KVVP'}
TWISTED_CORES = {'RYJSP', 'RYJS', 'RVS', 'RVSP'}
WIRE_CORES = {'BYJR', 'BYJ', 'BVR', 'BV'}
SOFT_CORES = {'RVV', 'RVVP'}

# PE/N 截面表
PE_TABLE = {
    2.5: 1.5, 4: 2.5, 6: 4, 10: 6, 16: 10, 25: 16, 35: 16,
    50: 25, 70: 35, 95: 50, 120: 70, 150: 70, 185: 95, 240: 120,
    300: 150, 400: 185
}


def normalize_model(raw):
    """标准化型号"""
    m = raw.strip()

    # 格式清理
    m = re.sub(r'[xX]', '×', m)      # x → ×
    m = re.sub(r'\*', '×', m)        # * → ×
    m = re.sub(r'\s+', '', m)        # 去空格
    m = re.sub(r'-+', '-', m)        # 多连字符→单
    m = re.sub(r'(\d)[kK][vV]', r'\1kV', m)  # kv→kV 统一
    m = m.strip('-')

    # 品牌型电缆特殊处理: NG-A 矿物绝缘
    is_ng_a = False
    if m.startswith('NG-A'):
        is_ng_a = True

    if is_ng_a:
        # NG-A: 直接提取电压和截面
        vm = re.search(r'(\d+\.?\d*/\d+\.?\d*kV|\d+kV)', m)
        voltage = vm.group(1) if vm else '0.6/1kV'
        # 提取截面: NG-A-0.6/1kv-4X70+1X35 → 4×70+1×35
        sec_match = re.search(r'(\d+)×(\d+\.?\d*)(?:\+(\d+)×(\d+\.?\d*))?', m)
        if not sec_match:
            return None
        cc = int(sec_match.group(1))
        sec = sec_match.group(2)
        pe = f'+{sec_match.group(3)}×{sec_match.group(4)}' if sec_match.group(3) and sec_match.group(4) else ''
        return f'NG-A-{voltage}-{cc}×{sec}{pe}'

    # ZA 前缀（阻燃A级）
    prefix = ''
    if re.match(r'ZA-', m):
        prefix = 'ZA'
        m = m[3:]

    # 提取核心型号
    core_found = ''
    for c in sorted(ALL_CORES, key=len, reverse=True):
        if m.startswith(c):
            core_found = c
            m = m[len(c):]
            break

    if not core_found:
        return None

    # 从剩余部分提取并移除电压
    voltage = ''
    vm = re.search(r'(\d+\.?\d*/\d+\.?\d*kV|\d+kV)', m)
    if vm:
        voltage = vm.group(1)
        m = m[:vm.start()] + m[vm.end():]  # 从 m 中移除电压部分
    if m.startswith('-'):
        m = m[1:]

    # 提取截面信息
    sec_match = re.search(r'(\d+)×(\d+\.?\d*)(?:\+(\d+)×(\d+\.?\d*))?', m)
    if sec_match:
        core_count = int(sec_match.group(1))
        section = sec_match.group(2)
        pe_section = f'+{sec_match.group(3)}×{sec_match.group(4)}' if sec_match.group(3) and sec_match.group(4) else ''
    else:
        bare = re.search(r'^(\d+\.?\d*)$', m)
        if bare:
            section = bare.group(1)
            core_count = 0
            pe_section = ''
        else:
            return None

    # 电压补全（如果原始没有电压）
    if not voltage:
        if core_found in POWER_CORES:
            voltage = '0.6/1kV'
        elif core_found in CONTROL_CORES:
            voltage = '450/750V'
        elif core_found in TWISTED_CORES:
            voltage = '300/300V'
        elif core_found in SOFT_CORES:
            voltage = '300/500V'
        elif core_found == 'BYJ' or core_found == 'BYJR':
            voltage = '450/750V'
        elif core_found == 'BVR':
            voltage = '450/750V'
        elif core_found == 'BV':
            s = float(section)
            voltage = '300/500V' if s <= 1.0 else '450/750V'
        else:
            voltage = '0.6/1kV'

    # 判断中压（≥6kV 豁免）
    is_mv = False
    kv_match = re.search(r'(\d+)kV', voltage)
    if kv_match and int(kv_match.group(1)) >= 6:
        is_mv = True

    # 构建标准型号
    sec_str = f'{core_count}×{section}' if core_count > 0 else f'1×{section}'
    if pe_section:
        sec_str += pe_section

    parts = []
    if prefix:
        parts.append(prefix)
    parts.append(core_found)
    if voltage:
        parts.append(voltage)
    parts.append(sec_str)

    std = '-'.join(parts)
    if is_mv:
        std = f'[MV]{std}'
    return std


def check_n_pe(std_model):
    """N/PE截面校验"""
    clean = std_model.replace('[MV]', '')
    warnings = []
    p = re.search(r'(\d+)×(\d+\.?\d*)\+(\d+)×(\d+\.?\d*)', clean)
    if p:
        ps = float(p.group(2))
        ns = float(p.group(4))
        for k in sorted(PE_TABLE.keys()):
            if ps <= k:
                if ns < PE_TABLE[k]:
                    warnings.append(
                        f'[PE报警] 相线{ps}mm²→PE需≥{PE_TABLE[k]}mm²，实配{ns}mm²'
                    )
                break
    return warnings


# ── 排序 ──────────────────────────────────────────────────────────────

def sort_key(model):
    if model.startswith('[MV]'):
        return (0, model)
    if 'NG-A' in model:
        return (1, model)
    if 'YJV22' in model:
        return (2, model)
    if 'YJV' in model or 'YJY' in model:
        return (3, model)
    if 'RVV' in model:
        return (4, model)
    return (5, model)


# ── 主流程 ────────────────────────────────────────────────────────────

def main():
    print('=' * 80)
    print('《原始数据索引》（共{}条）'.format(len(ALL_ENTRIES)))
    print('=' * 80)
    print(f'{"序号":>4s}  {"区域":<20s}  {"原始型号":<40s}  {"数量(m)":>10s}')
    print('-' * 80)

    for i, (model, qty) in enumerate(ALL_ENTRIES, 1):
        area = '机房楼外' if i <= len(OUTDOOR_CABLES) else '机房'
        print(f'{i:>4d}  {area:<20s}  {model:<40s}  {qty:>10.2f}')

    # 标准化
    print()
    print()
    print('=' * 80)
    print('逐条标准化及校验结果')
    print('=' * 80)

    groups = defaultdict(list)
    all_results = []

    for i, (model, qty) in enumerate(ALL_ENTRIES, 1):
        std = normalize_model(model)
        if std:
            warns = check_n_pe(std)
            groups[std].append((i, qty))
            all_results.append((i, model, std, warns))
        else:
            all_results.append((i, model, '【解析失败】', []))

    for i, orig, std, warns in all_results:
        ws = '; '.join(warns) if warns else ''
        print(f'{i:>4d}  {orig:<40s} → {std:<50s}  {ws}')

    # 按型号指纹分组合并
    print()
    print()
    print('=' * 80)
    print('按标准型号分组合并')
    print('=' * 80)

    sorted_groups = sorted(groups.items(), key=lambda x: sort_key(x[0]))

    grand_total = 0
    for std, entries in sorted_groups:
        total = sum(e[1] for e in entries)
        qty_strs = [f'#{idx}({qty}m)' for idx, qty in entries]
        print(f'\n{std}')
        print(f'  来源: {" + ".join(qty_strs)}')
        print(f'  合计: {total:.2f}m')
        grand_total += total

    # 标准化输出（TSV）
    print()
    print()
    print('=' * 80)
    print('标准化输出（TSV格式）')
    print('=' * 80)
    print('规格型号\t单位\t数量')

    for std, entries in sorted_groups:
        total = sum(e[1] for e in entries)
        total_rd = math.ceil(total * 100) / 100
        display = std.replace('[MV]', '')
        print(f'{display}\tm\t{total_rd:.2f}')

    print(f'\n原始{len(ALL_ENTRIES)}条 → 合并后{len(groups)}行')

    # 验算报告
    print()
    print()
    print('=' * 80)
    print('《合并验算报告》')
    print('=' * 80)
    print(f'{"规格型号":<55s} {"数量(m)":>12s}')
    print('-' * 70)
    grand = 0
    for std, entries in sorted_groups:
        total = sum(e[1] for e in entries)
        total_rd = math.ceil(total * 100) / 100
        display = std.replace('[MV]', '')
        print(f'{display:<55s} {total_rd:>12.2f}')
        grand += total_rd
    print('-' * 70)
    print(f'{"合计":<55s} {grand:>12.2f}')
    print(f'原始{len(ALL_ENTRIES)}条 → 合并后{len(groups)}行（合并{len(ALL_ENTRIES) - len(groups)}条）')

    # PE 报警
    pe_warns = [(i, orig, std, w) for i, orig, std, warns in all_results for w in warns]
    if pe_warns:
        print()
        print('⚠ PE截面报警：')
        for i, orig, std, w in pe_warns:
            print(f'  条目{i}: {orig:<40s} → {w}')

    return groups, sorted_groups, all_results


def export_excel(sorted_groups, all_results, output_path):
    """生成格式化 Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()

    # 样式
    hdr_font = Font(name='Microsoft YaHei', bold=True, size=11, color='FFFFFF')
    hdr_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    cat_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
    cat_font = Font(name='Microsoft YaHei', bold=True, size=11, color='2F5496')
    data_font = Font(name='Microsoft YaHei', size=10)
    total_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    total_font = Font(name='Microsoft YaHei', bold=True, size=11)
    thin = Side(style='thin', color='B4C6E7')
    bd = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center')
    left = Alignment(horizontal='left', vertical='center')

    # ── Sheet1: 标准化清单 ──
    ws1 = wb.active
    ws1.title = '标准化清单'

    headers = ['序号', '规格型号', '单位', '数量(m)']
    for c, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=c, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center
        cell.border = bd

    CAT_ORDER = [
        ('中压电力电缆 (YJV22-20kV)', lambda s: '[MV]' in s),
        ('矿物绝缘电缆 (NG-A)', lambda s: 'NG-A' in s),
        ('低压电力电缆 (YJV-0.6/1kV)', lambda s: 'YJV' in s),
        ('软电缆 (ZA-RVV)', lambda s: 'RVV' in s),
    ]

    row = 2
    seq = 1
    grand = 0

    for cat_name, cat_fn in CAT_ORDER:
        items = [(s, e) for s, e in sorted_groups if cat_fn(s)]
        if not items:
            continue

        # 类别标题行
        ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        cell = ws1.cell(row=row, column=1, value=cat_name)
        cell.font = cat_font
        cell.fill = cat_fill
        cell.alignment = Alignment(horizontal='left', vertical='center')
        for c in range(1, 5):
            ws1.cell(row=row, column=c).border = bd
        row += 1

        for std, entries in items:
            total = sum(e[1] for e in entries)
            total_rd = math.ceil(total * 100) / 100
            display = std.replace('[MV]', '')

            ws1.cell(row=row, column=1, value=seq).font = data_font
            ws1.cell(row=row, column=1).alignment = center
            ws1.cell(row=row, column=1).border = bd

            ws1.cell(row=row, column=2, value=display).font = data_font
            ws1.cell(row=row, column=2).alignment = left
            ws1.cell(row=row, column=2).border = bd

            ws1.cell(row=row, column=3, value='m').font = data_font
            ws1.cell(row=row, column=3).alignment = center
            ws1.cell(row=row, column=3).border = bd

            ws1.cell(row=row, column=4, value=total_rd).font = data_font
            ws1.cell(row=row, column=4).alignment = center
            ws1.cell(row=row, column=4).border = bd
            ws1.cell(row=row, column=4).number_format = '#,##0.00'

            row += 1
            seq += 1
            grand += total_rd

    # 合计行
    ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    cell = ws1.cell(row=row, column=1, value='合计')
    cell.font = total_font
    cell.fill = total_fill
    cell.alignment = Alignment(horizontal='right', vertical='center')
    for c in range(1, 5):
        ws1.cell(row=row, column=c).border = bd
        ws1.cell(row=row, column=c).fill = total_fill
    ws1.cell(row=row, column=4, value=round(grand, 2)).font = total_font
    ws1.cell(row=row, column=4).number_format = '#,##0.00'

    ws1.column_dimensions['A'].width = 8
    ws1.column_dimensions['B'].width = 48
    ws1.column_dimensions['C'].width = 8
    ws1.column_dimensions['D'].width = 15

    # ── Sheet2: 原始数据索引 ──
    ws2 = wb.create_sheet('原始数据索引')
    h2 = ['序号', '区域', '原始型号', '标准化型号', '数量(m)', 'PE校验']
    for c, h in enumerate(h2, 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center
        cell.border = bd

    for i, orig, std, warns in all_results:
        area = '机房楼外' if i <= len(OUTDOOR_CABLES) else '机房'
        ws_note = '; '.join(warns) if warns else 'OK'
        r = i + 1
        ws2.cell(row=r, column=1, value=i).font = data_font
        ws2.cell(row=r, column=2, value=area).font = data_font
        ws2.cell(row=r, column=3, value=orig).font = data_font
        display = std.replace('[MV]', '') if std else std
        ws2.cell(row=r, column=4, value=display).font = data_font
        ws2.cell(row=r, column=5, value=ALL_ENTRIES[i-1][1]).font = data_font
        ws2.cell(row=r, column=6, value=ws_note).font = data_font
        for c in range(1, 7):
            ws2.cell(row=r, column=c).border = bd
            ws2.cell(row=r, column=c).alignment = center if c != 3 else left

    ws2.column_dimensions['A'].width = 8
    ws2.column_dimensions['B'].width = 12
    ws2.column_dimensions['C'].width = 38
    ws2.column_dimensions['D'].width = 48
    ws2.column_dimensions['E'].width = 12
    ws2.column_dimensions['F'].width = 20

    wb.save(output_path)
    print(f'\nExcel文件已生成：{output_path}')


if __name__ == '__main__':
    groups, sorted_groups, all_results = main()
    output = 'C:/Users/yanli/Desktop/印尼数据中心项目_电缆审计结果.xlsx'
    export_excel(sorted_groups, all_results, output)
