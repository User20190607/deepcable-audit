#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
处理：项目材料需求表（电工材料）20260503 - 电线电缆.xls
"""

import re
import math
import xlrd
from collections import defaultdict

# ── 读取数据 ──────────────────────────────────────────────────────────

INPUT = r'C:\Users\yanli\Desktop\项目材料需求表（电工材料）20260503 - 电线电缆.xls'
OUTPUT = r'C:\Users\yanli\Desktop\项目材料需求表_审计结果.xlsx'

wb = xlrd.open_workbook(INPUT)
ws = wb.sheet_by_index(1)  # Sheet: 项目材料需求表

raw_entries = []
for r in range(3, ws.nrows):
    model = str(ws.cell_value(r, 3)).strip()
    buy_qty = ws.cell_value(r, 8)  # 本次申购数量
    if not model or model in ('', '型号规格', '楷模', '卡通'):
        continue
    try:
        qty = float(buy_qty) if buy_qty else 0.0
    except ValueError:
        continue
    if qty <= 0:
        continue
    raw_entries.append((model, qty))

# ── 规则定义 ──────────────────────────────────────────────────────────

ALL_CORES = ['YJV22', 'YJV', 'YJY', 'KYJYP', 'KYJY', 'RYJSP', 'RYJS', 'RYY',
             'RVS', 'RVSP', 'BYJR', 'BYJ', 'BVR', 'BV', 'BBTRZ', 'BTLY']

POWER_CORES = {'YJV22', 'YJV', 'YJY'}
CONTROL_CORES = {'KYJYP', 'KYJY'}
TWISTED_CORES = {'RYJSP', 'RYJS', 'RVS', 'RVSP'}
SOFT_CORES = {'RYY'}
SINGLE_CORE_WIRES = {'BYJ', 'BYJR', 'BVR', 'BV'}

def normalize_model(raw):
    m = raw.strip()

    # ── 颜色提取（最先，避免干扰型号解析） ──
    color = ''
    cm = re.search(r'[（(](\S+)[）)]$', m)
    if cm and cm.group(1) in ('红', '蓝', '黄', '绿', '双色', '白'):
        color = cm.group(1)
        m = m[:cm.start()].rstrip('-')

    # ── 平方清理 ──
    m = re.sub(r'平方', '', m)

    # ── 格式清理 ──
    m = re.sub(r'[xX]', '×', m)
    m = re.sub(r'\*', '×', m)
    m = re.sub(r'\s+', '', m)
    m = re.sub(r'(\d)[kK][vV]', r'\1kV', m)
    m = re.sub(r'-+', '-', m)
    m = m.strip('-')

    # ── 品牌/特殊型号预处理 ──
    is_btly = False
    b1_tag = ''
    if m.startswith('NW-BTLY'):
        is_btly = True
        b1m = re.search(r'B1[,，]t1[,，]d1', m)
        if b1m:
            b1_tag = 'B1(t1d1)'
            m = m[:b1m.start()] + m[b1m.end():]
        m = re.sub(r'-{2,}', '-', m).strip('-')

    # ── B1(t1d1) 通用处理（含括号格式） ──
    if not b1_tag:
        b1m = re.search(r'B1\(t1d1\)', m)
        if b1m:
            b1_tag = 'B1(t1d1)'
            m = m[:b1m.start()] + m[b1m.end():]
    m = re.sub(r'-{2,}', '-', m).strip('-')

    if is_btly:
        # NW-BTLY: 提取电压和截面
        vm = re.search(r'(\d+\.?\d*/\d+\.?\d*kV)', m)
        voltage = vm.group(1) if vm else '0.6/1kV'
        sec_match = re.search(r'(\d+)×(\d+\.?\d*)(?:\+(\d+)×(\d+\.?\d*))?', m)
        if not sec_match:
            return (None, 1)
        cc = int(sec_match.group(1))
        sec = sec_match.group(2)
        pe = f'+{sec_match.group(3)}×{sec_match.group(4)}' if sec_match.group(3) and sec_match.group(4) else ''
        result = f'NW-BTLY-{voltage}-{cc}×{sec}{pe}'
        if b1_tag:
            result = f'NW-BTLY-{b1_tag}-{voltage}-{cc}×{sec}{pe}'
        if color:
            result += f'({color})'
        return (result, 1)

    # ── ZR- / ZN- 前缀处理 ──
    if m.startswith('ZR-'):
        m = 'ZC-' + m[3:]
    elif m.startswith('ZN-'):
        m = 'ZCN-' + m[3:]

    # ── 提取前缀 ──
    all_prefixes = sorted(['WDZBN', 'WDZCN', 'WDZN', 'WDZB', 'WDZC', 'WDZ', 'ZC', 'ZCN'], key=len, reverse=True)
    prefix = ''
    for p in all_prefixes:
        if m.startswith(p):
            prefix = p
            m = m[len(p):]
            if m.startswith('-'):
                m = m[1:]
            break

    # ── 提取核心型号 ──
    core_found = ''
    core_starts = sorted(ALL_CORES, key=len, reverse=True)
    # 特殊处理 BTLY（无前缀时）
    if m.startswith('BTLY'):
        core_found = 'BTLY'
        m = m[5:]
    else:
        for c in core_starts:
            if m.startswith(c):
                core_found = c
                m = m[len(c):]
                break
    if not core_found:
        return (None, 1)
    if m.startswith('-'):
        m = m[1:]

    # ── 电压提取 ──
    voltage = ''
    vm = re.search(r'(\d+\.?\d*/\d+\.?\d*kV|\d+kV)', m)
    if vm:
        voltage = vm.group(1)
        m = m[:vm.start()] + m[vm.end():]
    if m.startswith('-'):
        m = m[1:]

    # ── B1 标记（-B1- 或 B1- 前置，不含括号格式） ──
    b1_dash = ''
    b1m = re.match(r'B1-?', m)
    if b1m:
        b1_dash = 'B1'
        m = m[b1m.end():]
        if m.startswith('-'):
            m = m[1:]

    # ── 截面提取 ──
    sec_match = re.search(r'(\d+)×(\d+\.?\d*)(?:\+(\d+)×(\d+\.?\d*))?', m)
    if sec_match:
        core_count = int(sec_match.group(1))
        section = sec_match.group(2)
        pe_section = f'+{sec_match.group(3)}×{sec_match.group(4)}' if sec_match.group(3) and sec_match.group(4) else ''
    else:
        bare = re.search(r'^(\d+\.?\d*)$', m)
        if bare:
            section = bare.group(1)
            core_count = 0
            pe_section = ''
        else:
            return (None, 1)

    # ── §6.3 单芯线拆分：BYJ/BV/BVR/BYJR 多芯写法拆为 1×，数量×N ──
    split_factor = 1
    if core_found in SINGLE_CORE_WIRES and core_count > 1:
        split_factor = core_count
        core_count = 1
        pe_section = ''

    # ── 电压补全 ──
    if not voltage:
        if core_found in POWER_CORES:
            voltage = '0.6/1kV'
        elif core_found in CONTROL_CORES:
            voltage = '450/750V'
        elif core_found in TWISTED_CORES:
            voltage = '300/300V'
        elif core_found in SOFT_CORES:
            voltage = '300/500V'
        elif core_found == 'BYJ' or core_found == 'BYJR':
            voltage = '450/750V'
        elif core_found == 'BVR':
            voltage = '450/750V'
        elif core_found == 'BV':
            s = float(section)
            voltage = '300/500V' if s <= 1.0 else '450/750V'
        else:
            voltage = '0.6/1kV'

    # ── 中压判断 ──
    is_mv = False
    kv_match = re.search(r'(\d+)kV', voltage)
    if kv_match and int(kv_match.group(1)) >= 6:
        is_mv = True

    # ── 材质互斥: WD + V 修正 ──
    # ZC-YJV → ZC prefix, YJV core. Has V but no WD → 不触发
    if prefix.startswith('WD') and core_found == 'YJV':
        core_found = 'YJY'

    # ── 构建 ──
    sec_str = f'{core_count}×{section}' if core_count > 0 else f'1×{section}'
    if pe_section:
        sec_str += pe_section

    parts = []
    if prefix:
        parts.append(prefix)
    parts.append(core_found)
    if b1_tag:
        parts.append(b1_tag)
    elif b1_dash:
        parts.append(b1_dash)
    if voltage:
        parts.append(voltage)
    parts.append(sec_str)

    std = '-'.join(parts)
    if is_mv:
        std = f'[MV]{std}'
    if color:
        std += f'({color})'
    return (std, split_factor)


# ── N/PE 校验 ─────────────────────────────────────────────────────────

PE_TABLE = {
    2.5: 1.5, 4: 2.5, 6: 4, 10: 6, 16: 10, 25: 16, 35: 16,
    50: 25, 70: 35, 95: 50, 120: 70, 150: 70, 185: 95, 240: 120,
    300: 150, 400: 185
}

def check_n_pe(std):
    clean = std.replace('[MV]', '')
    p = re.search(r'(\d+)×(\d+\.?\d*)\+(\d+)×(\d+\.?\d*)', clean)
    if not p:
        return []
    ps = float(p.group(2))
    ns = float(p.group(4))
    for k in sorted(PE_TABLE):
        if ps <= k:
            if ns < PE_TABLE[k]:
                return [f'[PE报警] 相线{ps}mm²→PE需≥{PE_TABLE[k]}mm²，实配{ns}mm²']
            break
    return []


# ── 排序 ──────────────────────────────────────────────────────────────

def sort_key(std):
    if '[MV]' in std: return (0, std)
    if 'NW-BTLY' in std: return (1, std)
    if 'YJV' in std or 'YJY' in std: return (2, std)
    if 'KYJY' in std: return (3, std)
    if 'BYJ' in std or 'BYJR' in std: return (4, std)
    if 'RYJS' in std or 'RYJSP' in std or 'RYY' in std: return (5, std)
    return (6, std)


# ── 主流程 ────────────────────────────────────────────────────────────

print('=' * 90)
print(f'《原始数据索引》（共{len(raw_entries)}条）')
print('=' * 90)
print(f'{"序号":>4s}  {"原始型号":<45s}  {"申购数量":>10s}')
print('-' * 65)
for i, (m, q) in enumerate(raw_entries, 1):
    print(f'{i:>4d}  {m:<45s}  {q:>10.2f}')

groups = defaultdict(list)
results = []

for i, (model, qty) in enumerate(raw_entries, 1):
    std, split_factor = normalize_model(model)
    effective_qty = qty * split_factor
    warns = check_n_pe(std) if std else []
    if std:
        groups[std].append((i, effective_qty))
    results.append((i, model, effective_qty, std, warns))

print()
print('=' * 90)
print('逐条标准化及校验结果')
print('=' * 90)
for i, orig, eqty, std, warns in results:
    display = std if std else '【解析失败】'
    ws = '; '.join(warns) if warns else ''
    print(f'{i:>4d}  {orig:<42s} → {display:<55s}  {ws}')

sorted_groups = sorted(groups.items(), key=lambda x: sort_key(x[0]))

print()
print('=' * 90)
print('分组合并')
print('=' * 90)

grand = 0
for std, entries in sorted_groups:
    total = sum(e[1] for e in entries)
    detail = ' + '.join(f'#{idx}({qty}m)' for idx, qty in entries)
    print(f'\n{std}')
    print(f'  来源: {detail}')
    print(f'  合计: {total:.2f}m')
    grand += total

print()
print('=' * 90)
print('标准化输出（TSV）')
print('=' * 90)
print('规格型号\t单位\t数量')
for std, entries in sorted_groups:
    total = sum(e[1] for e in entries)
    display = std.replace('[MV]', '')
    print(f'{display}\tm\t{total:.2f}')

print(f'\n原始{len(raw_entries)}条 → 合并后{len(groups)}行')
print(f'合计总长度: {grand:.2f}m')

# PE 报警
pe_warns = [(i, orig, std, w) for i, orig, eqty, std, warns in results for w in warns]
if pe_warns:
    print()
    print('⚠ PE截面报警：')
    for i, orig, std, w in pe_warns:
        print(f'  条目{i}: {orig:<42s} → {w}')

# ── Excel 输出 ──
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb2 = Workbook()
hdr_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
hdr_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
cat_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
cat_font = Font(name='Arial', bold=True, size=11, color='2F5496')
data_font = Font(name='Arial', size=10)
total_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
total_font = Font(name='Arial', bold=True, size=11)
thin = Side(style='thin', color='B4C6E7')
bd = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal='center', vertical='center')
left = Alignment(horizontal='left', vertical='center')

# Sheet1: 标准化清单
ws1 = wb2.active
ws1.title = '标准化清单'
headers = ['序号', '规格型号', '单位', '数量(m)']
for c, h in enumerate(headers, 1):
    cell = ws1.cell(row=1, column=c, value=h)
    cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = center; cell.border = bd

CATS = [
    ('矿物绝缘电缆 (NW-BTLY)', lambda s: 'NW-BTLY' in s),
    ('低烟无卤电力电缆 (WDZB-YJY)', lambda s: 'YJY' in s),
    ('阻燃电力电缆 (ZC-ZCN-YJV)', lambda s: 'YJV' in s and 'ZC' in s),
    ('控制电缆 (KYJY)', lambda s: 'KYJY' in s),
    ('无卤布电线 (WDZC-BYJ)', lambda s: 'BYJ' in s),
    ('双绞线/屏蔽线 (RYJS/RYJSP/RYY)', lambda s: any(x in s for x in ['RYJS', 'RYJSP', 'RYY'])),
]

row = 2; seq = 1; grand_total = 0
for cat_name, cat_fn in CATS:
    items = [(s, e) for s, e in sorted_groups if cat_fn(s)]
    if not items: continue
    ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    cell = ws1.cell(row=row, column=1, value=cat_name)
    cell.font = cat_font; cell.fill = cat_fill; cell.alignment = Alignment(horizontal='left', vertical='center')
    for c in range(1, 5): ws1.cell(row=row, column=c).border = bd
    row += 1
    for std, entries in items:
        total = sum(e[1] for e in entries)
        display = std.replace('[MV]', '')
        ws1.cell(row=row, column=1, value=seq).font = data_font; ws1.cell(row=row, column=1).alignment = center; ws1.cell(row=row, column=1).border = bd
        ws1.cell(row=row, column=2, value=display).font = data_font; ws1.cell(row=row, column=2).alignment = left; ws1.cell(row=row, column=2).border = bd
        ws1.cell(row=row, column=3, value='m').font = data_font; ws1.cell(row=row, column=3).alignment = center; ws1.cell(row=row, column=3).border = bd
        ws1.cell(row=row, column=4, value=total).font = data_font; ws1.cell(row=row, column=4).alignment = center; ws1.cell(row=row, column=4).border = bd
        row += 1; seq += 1; grand_total += total

ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
cell = ws1.cell(row=row, column=1, value='合计')
cell.font = total_font; cell.fill = total_fill; cell.alignment = Alignment(horizontal='right', vertical='center')
for c in range(1, 5): ws1.cell(row=row, column=c).border = bd; ws1.cell(row=row, column=c).fill = total_fill
ws1.cell(row=row, column=4, value=round(grand_total, 2)).font = total_font
ws1.column_dimensions['A'].width = 8
ws1.column_dimensions['B'].width = 55
ws1.column_dimensions['C'].width = 8
ws1.column_dimensions['D'].width = 15

# Sheet2: 原始数据索引
ws2 = wb2.create_sheet('原始数据索引')
h2 = ['序号', '原始型号', '标准化型号', '申购数量(m)', 'PE校验']
for c, h in enumerate(h2, 1):
    cell = ws2.cell(row=1, column=c, value=h)
    cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = center; cell.border = bd

for i, orig, eqty, std, warns in results:
    r = i + 1
    display = std.replace('[MV]', '') if std else '解析失败'
    ws_note = '; '.join(warns) if warns else 'OK'
    vals = [i, orig, display, eqty, ws_note]
    for c, v in enumerate(vals, 1):
        cell = ws2.cell(row=r, column=c, value=v)
        cell.font = data_font; cell.border = bd
        cell.alignment = left if c in (2, 3) else center

ws2.column_dimensions['A'].width = 8
ws2.column_dimensions['B'].width = 45
ws2.column_dimensions['C'].width = 55
ws2.column_dimensions['D'].width = 14
ws2.column_dimensions['E'].width = 20

wb2.save(OUTPUT)
print(f'\nExcel: {OUTPUT}')
