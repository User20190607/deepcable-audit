import sys, os, re, math
sys.path.insert(0, r'C:\Users\yanli\.claude\commands\deepcable-audit')
from deepcable_normalize import normalize

def extract_model(text):
    """Extract cable model from multi-line description text."""
    s = str(text)
    # Try 型号: field first
    m = re.search(r'型号[:：]\s*([A-Za-z0-9\-\*\s\.\+\/×]+?)(?:[\n\\n]|$)', s)
    if m:
        return m.group(1).strip()
    # Try 名称: field with 电线/电缆
    m = re.search(r'(?:电线|电缆)\s*([A-Za-z0-9\-\s\.\+/]+?)(?:[\n\\n]|$)', s)
    if m:
        return m.group(1).strip()
    return None

def clean_model(raw):
    """Preprocess model string for normalize()."""
    if not raw:
        return None
    s = raw.strip()
    # Handle non-standard WDNH prefix
    s = re.sub(r'^WDNH-', 'WDZN-', s)
    s = re.sub(r'^WDNH(?![A-Z])', 'WDZN', s)
    # Convert 750 voltage abbreviation between model name and section
    s = re.sub(r'\b750[\s-]*(?=\d)', '450/750V-', s)
    # Strip mm2 units
    s = re.sub(r'[Mm][Mm][2²]?', '', s)
    s = re.sub(r'平方', '', s)
    s = re.sub(r'\s+', '', s)
    return s

def process(input_path, output_path):
    import pandas as pd
    df = pd.read_excel(input_path, header=None)

    rows = []
    failed = []
    for i in range(2, len(df)):  # skip title + header
        seq = df.iloc[i, 0]
        raw_text = str(df.iloc[i, 2]) if pd.notna(df.iloc[i, 2]) else ''
        qty = df.iloc[i, 4] if df.shape[1] > 4 and pd.notna(df.iloc[i, 4]) else None

        ext = extract_model(raw_text)
        if not ext:
            failed.append((seq, raw_text[:60], '未提取到型号'))
            continue

        cleaned = clean_model(ext)
        if not cleaned:
            failed.append((seq, ext, '清理后为空'))
            continue

        std = normalize(cleaned)
        if std:
            rows.append({'seq': seq, 'raw': ext, 'std': std, 'qty': qty})
        else:
            failed.append((seq, cleaned, '标准化失败'))

    # Output with xlsx skill patterns
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    wb = Workbook(); ws = wb.active
    ws.title = '标准化清单'
    thin = Side(style='thin', color='CCCCCC')
    bd = Border(left=thin, right=thin, top=thin, bottom=thin)
    hf = PatternFill('solid', fgColor='1A3A6B')
    hdr_f = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    df = Font(name='Arial', size=10)
    ca = Alignment(horizontal='center', vertical='center')
    la = Alignment(horizontal='left', vertical='center')
    for col, (h, w) in enumerate(zip(['序号', '原始型号', '标准型号', '单位'], [8, 60, 60, 8]), 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = hdr_f; c.fill = hf; c.alignment = ca; c.border = bd
        ws.column_dimensions[get_column_letter(col)].width = w
    alt = PatternFill('solid', fgColor='EEF4FB')
    for i, r in enumerate(rows, 2):
        fill = alt if i % 2 == 0 else PatternFill('solid', fgColor='FFFFFF')
        for col, v in enumerate([r['seq'], r['raw'], r['std'], 'm'], 1):
            c = ws.cell(row=i, column=col, value=v)
            c.fill = fill; c.border = bd; c.font = df
            c.alignment = ca if col in (1, 4) else la
    tr = len(rows) + 2
    tf = Font(name='Arial', bold=True, size=10)
    tfill = PatternFill('solid', fgColor='D6E4F0')
    for col, v in enumerate([None, '合计', None, f'=SUM(D2:D{tr-1})'], 1):
        c = ws.cell(row=tr, column=col, value=v)
        c.font = tf; c.fill = tfill; c.border = bd; c.alignment = ca
    ws.freeze_panes = 'A2'

    # Failed items sheet
    if failed:
        ws2 = wb.create_sheet('解析失败')
        ws2.cell(row=1, column=1, value='序号').font = hdr_f
        ws2.cell(row=1, column=1).fill = hf
        ws2.cell(row=1, column=2, value='原始内容').font = hdr_f
        ws2.cell(row=1, column=2).fill = hf
        ws2.cell(row=1, column=3, value='原因').font = hdr_f
        ws2.cell(row=1, column=3).fill = hf
        for i, (seq, raw_short, reason) in enumerate(failed, 2):
            ws2.cell(row=i, column=1, value=seq)
            ws2.cell(row=i, column=2, value=raw_short)
            ws2.cell(row=i, column=3, value=reason)

    wb.save(output_path)
    return len(rows), len(failed)

if __name__ == '__main__':
    n, f = process(
        r'C:\Users\yanli\Desktop\电缆、电线(2).xls',
        r'C:\Users\yanli\Desktop\电线电缆标准化清单.xlsx'
    )
    print(f'OK: {n} 条, 失败: {f} 条')
