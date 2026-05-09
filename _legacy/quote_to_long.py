"""
报价表长表转换脚本 v2
宽表 → 长表 → 去重（同一型号保留最低价）
"""
import pandas as pd
import sys, os, math
from collections import defaultdict

sys.path.insert(0, os.path.expanduser('~/.claude/commands/deepcable-audit'))
from deepcable_normalize import normalize

INPUT  = os.path.expanduser('~/Desktop/2025电线电缆（贸易商）报价（中海领潮）2.xlsx')
OUTPUT = os.path.expanduser('~/Desktop/报价长表_标准化.xlsx')

# 阻燃等级列 → 前缀映射
COL_PREFIXES = {
    6: '', 7: 'NH-', 8: 'ZA-', 9: 'ZAN-', 10: 'ZB-', 11: 'ZBN-',
    12: 'ZC-', 13: 'ZCN-', 14: 'WDZA-', 15: 'WDZB-',
    16: 'WDZB1-', 17: 'WDZB2-', 18: 'WDZC-', 19: 'WDZAN-', 20: 'WDZBN-', 21: 'WDZCN-',
    22: 'WDZCN-B1-', 23: 'WDZCN-B2-', 24: 'WDZA-B1-', 25: 'WDZB-B1-',
    26: 'WDZAN-B1-', 27: 'WDZBN-B1-', 28: 'WDZAN-B2-', 29: 'WDZBN-B2-'
}

df = pd.read_excel(INPUT, sheet_name=2, header=None, dtype=str)

# 用 dict 按标准型号聚合，保留最低总价
# {标准型号: {'总价': min_price, '铜净重': w, '加工费': fee, '来源数': count}}
best = {}
skip_no_price = 0
skip_invalid = 0

for i in range(1, len(df)):  # ← 从 row 1 开始（row 0 是表头）
    base_model = str(df.iloc[i, 2]).strip()
    if not base_model or base_model in ('nan', 'None', ''):
        continue

    try:
        copper_price = float(str(df.iloc[i, 5]).replace(',', ''))
    except (ValueError, TypeError):
        copper_price = 0.0
    try:
        copper_weight = float(str(df.iloc[i, 4]).replace(',', ''))
    except (ValueError, TypeError):
        copper_weight = 0.0

    for col, prefix in COL_PREFIXES.items():
        raw_val = df.iloc[i, col]
        val_str = str(raw_val).strip() if pd.notna(raw_val) else ''
        if not val_str or val_str in ('—', '-', 'nan', 'None', ''):
            skip_no_price += 1
            continue
        try:
            fee = float(val_str)
        except ValueError:
            skip_no_price += 1
            continue

        full_model = prefix + base_model
        std = normalize(full_model)
        if not std:
            skip_invalid += 1
            continue

        total_price = copper_price + fee

        if std not in best or total_price < best[std]['总价']:
            best[std] = {
                '总价': round(total_price, 4),
                '铜净重(kg/m)': copper_weight,
                '加工费(元/米)': fee,
                '原始型号': full_model,
            }

print(f'去重后标准型号: {len(best)}')
print(f'跳过(无价格): {skip_no_price}')
print(f'跳过(标准化失败): {skip_invalid}')

# 排序：电力 → 控制 → 布电线 → 软线
def sort_key(m):
    if '10KV' in m or '35KV' in m: return (0, m)
    if 'YJV' in m or 'YJY' in m:   return (1, m)
    if 'BBTRZ' in m:               return (2, m)
    if 'KYJY' in m:                return (3, m)
    if 'BVR' in m or 'BV' in m or 'BYJ' in m or 'BYJR' in m: return (4, m)
    if 'RYJ' in m or 'RVS' in m:   return (5, m)
    return (6, m)

sorted_models = sorted(best.items(), key=lambda x: sort_key(x[0]))

# 输出 Excel
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()
ws = wb.active
ws.title = '报价长表'

headers = ['序号', '标准型号', '总价(元/米)', '铜净重(kg/m)', '加工费(元/米)']
widths  = [8, 60, 14, 14, 14]
thin = Side(style='thin', color='CCCCCC')
bd = Border(left=thin, right=thin, top=thin, bottom=thin)
hdr_fill = PatternFill('solid', fgColor='1A3A6B')
hdr_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
data_font = Font(name='Arial', size=10)
center = Alignment(horizontal='center', vertical='center')
left   = Alignment(horizontal='left', vertical='center')

for c, (h, w) in enumerate(zip(headers, widths), 1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.alignment = center
    cell.border = bd
    ws.column_dimensions[chr(64+c)].width = w

for idx, (model, info) in enumerate(sorted_models, 1):
    r = idx + 1
    vals = [idx, model, info['总价'], info['铜净重(kg/m)'], info['加工费(元/米)']]
    for c, v in enumerate(vals, 1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = data_font
        cell.border = bd
        cell.alignment = center if c != 2 else left

# 合计行
total_row = len(sorted_models) + 2
ws.cell(row=total_row, column=1, value='合计').font = Font(name='Arial', bold=True, size=10)
ws.cell(row=total_row, column=1).alignment = Alignment(horizontal='right', vertical='center')
ws.cell(row=total_row, column=3, value=f'=SUM(C2:C{total_row-1})').font = Font(name='Arial', bold=True, size=10)
for c in range(1, 6):
    ws.cell(row=total_row, column=c).border = bd
    ws.cell(row=total_row, column=c).fill = PatternFill('solid', fgColor='D6E4F0')

ws.freeze_panes = 'A2'
wb.save(OUTPUT)
print(f'\n输出: {OUTPUT}')
print(f'共 {len(best)} 条唯一标准型号报价')

# 校验：和 Sheet2 对比
print('\n=== 抽查对比（Sheet2 总价 vs 计算值） ===')
df2 = pd.read_excel(INPUT, sheet_name=1, header=None, dtype=str, nrows=10)
samples = [
    ('BV-450/750V-1.5',  '普通', ''),   # BV
    ('BV-450/750V-1.5',  'NH',   'NH-'),
    ('BV-450/750V-2.5',  '普通', ''),
    ('BV-450/750V-2.5',  'NH',   'NH-'),
]
for base_model, level_name, prefix in samples:
    # Sheet2 校验
    if base_model == 'BV-450/750V-1.5':
        s2_val = df2.iloc[3, 3]  # 普通
        if level_name == 'NH':
            s2_val = df2.iloc[3, 4]
    elif base_model == 'BV-450/750V-2.5':
        s2_val = df2.iloc[4, 3]
        if level_name == 'NH':
            s2_val = df2.iloc[4, 4]

    std_key = normalize(prefix + base_model) if prefix else normalize(base_model)
    our_val = best.get(std_key, {}).get('总价', 'N/A')
    print(f'  {std_key:<55s} Sheet2={s2_val}  ours={our_val}')
