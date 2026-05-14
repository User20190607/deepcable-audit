"""
excel_output.py — 6-Sheet 审计输出

Sheet 结构：
  1. 原始清单       — 客户文件原样复制，冻结保护
  2. 电缆提取+标准化 — 原始 vs 标准化，颜色编码
  3. 聚合采购清单    — 对外交付，分组+验算+追溯
  4. 型号解释        — 面向采购人员的通俗说明
  5. 待处理项        — 需要人工介入的问题汇总
  6. 操作日志        — 审计追溯

向后兼容：
    to_excel(aggregated, path) 包装为简化输出
"""

import re
import math
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# ── 颜色常量 ──
HDR_FILL     = PatternFill('solid', fgColor='1A3A6B')
HDR_FONT     = Font(name='Arial', bold=True, color='FFFFFF', size=11)
BODY_FONT    = Font(name='Arial', size=10)
BOLD_FONT    = Font(name='Arial', bold=True, size=10)
LINK_FONT    = Font(name='Arial', size=10, color='2E75B6', underline='single')

FILL_WHITE   = PatternFill('solid', fgColor='FFFFFF')
FILL_ALT     = PatternFill('solid', fgColor='F2F7FB')   # subtle blue
FILL_ORANGE  = PatternFill('solid', fgColor='FFF0E0')   # 有修正
FILL_GREEN   = PatternFill('solid', fgColor='E8F5E9')   # 无修正
FILL_GREY    = PatternFill('solid', fgColor='F0F0F0')   # 已排除
FILL_WARN    = PatternFill('solid', fgColor='FDECEA')   # 警告
FILL_TOTAL   = PatternFill('solid', fgColor='D6E4F0')
FILL_CAT     = PatternFill('solid', fgColor='E8EEF5')   # 分类标题

THIN_SIDE    = Side(style='thin', color='D0D0D0')
BORDER       = Border(left=THIN_SIDE, right=THIN_SIDE,
                      top=THIN_SIDE, bottom=THIN_SIDE)

CENTER       = Alignment(horizontal='center', vertical='center')
LEFT         = Alignment(horizontal='left', vertical='center', wrap_text=True)
LEFT_NOWRAP  = Alignment(horizontal='left', vertical='center')


def build_output(*, records, model_results, aggregated, issues, log_entries,
                 original_path, output_path, project_name=''):
    """
    生成 6-Sheet 审计 Excel。

    参数：
        records:       read_cable_list 原始输出（含非电缆行）
        model_results: [{raw, model, qty, change_log, change_type, pe_status, is_cable, row, sheet}, ...]
        aggregated:    [{model, qty, source_rows, verification_expr}, ...]
        issues:        [{type, raw_spec, description, suggestion}, ...]
        log_entries:   [{phase, item_ref, operation, rule}, ...]
        original_path: 原始文件路径（用于 Sheet1 复制）
        output_path:   输出文件路径
        project_name:  项目名（用于文件命名）
    """
    wb = Workbook()

    _build_sheet1(wb, original_path)
    _build_sheet2(wb, model_results)
    _build_sheet3(wb, aggregated)
    _build_sheet4(wb, aggregated)
    _build_sheet5(wb, issues)
    _build_sheet6(wb, log_entries)

    wb.save(output_path)
    return output_path


# ════════════════════════════════════════════════════════════════
# Sheet1  原始清单
# ════════════════════════════════════════════════════════════════

def _build_sheet1(wb, original_path):
    """完整复制原始工作表到 Sheet1，保留所有格式（合并单元格、列宽、行高、字体、填充、边框、对齐、数字格式）"""
    ws = wb.active
    ws.title = '原始清单'

    src = load_workbook(original_path)
    src_ws = src.active

    # ── 复制列宽 ──
    for col_letter, dim in src_ws.column_dimensions.items():
        ws.column_dimensions[col_letter].width = dim.width

    # ── 复制行高 ──
    for row_num, dim in src_ws.row_dimensions.items():
        ws.row_dimensions[row_num].height = dim.height

    # ── 复制合并单元格范围 ──
    merge_ranges = [str(m) for m in src_ws.merged_cells.ranges]

    # ── 复制所有单元格（值 + 格式），跳过合并单元格中的非顶级单元格 ──
    for row in src_ws.iter_rows(min_row=1, max_row=src_ws.max_row,
                                max_col=src_ws.max_column):
        for cell in row:
            # openpyxl's iter_rows skips MergedCell, but load_workbook returns them
            # Only write to non-merged cells; merged ranges are handled separately
            try:
                new_cell = ws.cell(row=cell.row, column=cell.column)
                new_cell.value = cell.value
            except AttributeError:
                continue  # MergedCell — skipped

            # 复制样式
            if not cell.has_style:
                continue
            try:
                new_cell.font = Font(
                    name=cell.font.name, size=cell.font.size,
                    bold=cell.font.bold, italic=cell.font.italic,
                    color=cell.font.color, underline=cell.font.underline,
                    strike=cell.font.strike,
                )
            except Exception:
                pass
            try:
                fg = cell.fill.fgColor
                if cell.fill.patternType:
                    fg_rgb = fg.rgb if fg and fg.rgb and fg.rgb != '00000000' else None
                    new_cell.fill = PatternFill('solid', fgColor=fg_rgb or 'FFFFFF')
            except Exception:
                pass
            try:
                bl = cell.border.left
                br = cell.border.right
                bt = cell.border.top
                bb = cell.border.bottom
                new_cell.border = Border(
                    left=Side(style=bl.style, color=bl.color) if bl and bl.style else None,
                    right=Side(style=br.style, color=br.color) if br and br.style else None,
                    top=Side(style=bt.style, color=bt.color) if bt and bt.style else None,
                    bottom=Side(style=bb.style, color=bb.color) if bb and bb.style else None,
                )
            except Exception:
                pass
            try:
                new_cell.alignment = Alignment(
                    horizontal=cell.alignment.horizontal,
                    vertical=cell.alignment.vertical,
                    wrap_text=cell.alignment.wrap_text,
                )
            except Exception:
                pass
            try:
                if cell.number_format and cell.number_format != 'General':
                    new_cell.number_format = cell.number_format
            except Exception:
                pass

    # ── 设置合并单元格 ──
    for mr in merge_ranges:
        ws.merge_cells(mr)

    src.close()

    # ── 插入页眉行 ──
    ws.insert_rows(1)

    # 找出实际有数据的最大列数（避免空列导致合并过宽）
    max_data_col = 1
    for row in src_ws.iter_rows(min_row=1, max_row=src_ws.max_row):
        for cell in row:
            if cell.value is not None:
                max_data_col = max(max_data_col, cell.column)
    max_data_col = max(max_data_col, 6)  # 至少6列

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_data_col)
    src_name = Path(original_path).name
    hdr = ws.cell(row=1, column=1,
                  value=f'原始文件：{src_name} | 读取日期：{datetime.now().strftime("%Y-%m-%d %H:%M")}')
    hdr.font = Font(name='Arial', size=9, color='666666', italic=True)
    ws.row_dimensions[1].height = 22

    # ── 冻结原始表头行（页眉下第一行） ──
    ws.freeze_panes = 'A3'

    # ── 保护工作表 ──
    ws.protection.sheet = True


# ════════════════════════════════════════════════════════════════
# Sheet2  电缆提取+标准化
# ════════════════════════════════════════════════════════════════

def _build_sheet2(wb, model_results):
    """核心工作表：原始 vs 标准化，颜色编码"""
    ws = wb.create_sheet('电缆提取+标准化')

    headers = [
        ('原始行号', 10), ('原始型号规格', 42), ('原始数量', 10), ('原始单位', 8),
        ('── 分隔线 ──', 4),
        ('标准化型号', 44), ('换算数量(m)', 12), ('修正类型', 14), ('修正说明', 32),
        ('PE校验', 12), ('处理状态', 12),
    ]
    _write_header_row(ws, headers, row=1)

    for i, mr in enumerate(model_results, start=2):
        is_cable = mr.get('is_cable', True)
        model   = mr.get('model', '')
        status  = mr.get('status', '✅正常')
        pe      = mr.get('pe_status', '—')
        ctype   = mr.get('change_type', '无修正')
        cdesc   = mr.get('change_desc', '')
        raw_s   = mr.get('raw', '')
        raw_q   = mr.get('raw_qty', 0)
        raw_u   = mr.get('raw_unit', 'm')
        norm_q  = mr.get('qty', 0)
        row_ref = mr.get('row', '')

        # Determine fill color
        if not is_cable:
            fill = FILL_GREY
        elif not model or status == '❌失败':
            fill = FILL_GREY
        elif ctype == '无修正':
            fill = FILL_GREEN
        elif status == '⚠️警告':
            fill = FILL_WARN
        else:
            fill = FILL_ORANGE

        # 原始行号 → 超链接到 Sheet1
        link_to_sheet1(ws, row=i, col=1, target_row=row_ref, display=str(row_ref))
        _cell(ws, i, 2, raw_s, fill, LEFT)
        _cell(ws, i, 3, _fmt_qty(raw_q), fill, CENTER)
        _cell(ws, i, 4, raw_u, fill, CENTER)
        _cell(ws, i, 5, '──', fill, CENTER)

        if is_cable and model:
            _cell(ws, i, 6, model, fill, LEFT)
            _cell(ws, i, 7, _fmt_qty(norm_q), fill, CENTER)
            _cell(ws, i, 8, ctype, fill, CENTER)
            _cell(ws, i, 9, cdesc, fill, LEFT)
            _cell(ws, i, 10, pe, fill, CENTER)
            _cell(ws, i, 11, status, fill, CENTER)

            # 颜色标记：PE 警告红色字体
            if '⚠' in pe:
                ws.cell(row=i, column=10).font = Font(name='Arial', size=10, color='C0392B', bold=True)
        else:
            _cell(ws, i, 6, model or '（已排除-非电线电缆）', fill, LEFT)
            for c in range(7, 12):
                _cell(ws, i, c, '—', fill, CENTER)

        ws.row_dimensions[i].height = 20

    ws.freeze_panes = 'A2'
    _set_auto_filter(ws, len(model_results) + 1, 11)


# ════════════════════════════════════════════════════════════════
# Sheet3  聚合采购清单
# ════════════════════════════════════════════════════════════════

def _build_sheet3(wb, aggregated):
    """对外采购表：分类分组 + 数量验算 + 来源追溯"""
    ws = wb.create_sheet('聚合采购清单')

    headers = [
        ('序号', 8), ('规格型号（标准化）', 48), ('单位', 6),
        ('数量', 12), ('型号解释（简版）', 40),
        ('来源行', 24), ('数量验算式', 28), ('备注', 20),
    ]
    _write_header_row(ws, headers, row=1)

    # 分类顺序
    CATEGORIES = ['中压电力电缆', '低压电力电缆', '控制电缆', '布电线', '软线/弱电线', '其他']
    def _categorize(model):
        if '10kV' in model or '35kV' in model: return '中压电力电缆'
        if any(x in model for x in ['KVV', 'KVVP', 'KYJY', 'KYJYP', 'DJYPV', 'DJYP']): return '控制电缆'
        if any(x in model for x in ['BV', 'BVR', 'BYJ', 'BYJR', 'BLV', 'BX', 'BLX']): return '布电线'
        if any(x in model for x in ['RVV', 'RVVP', 'RVS', 'RVB', 'RYJS', 'RYJSP', 'RYSP', 'RS485']): return '软线/弱电线'
        if any(x in model for x in ['YJV', 'YJY', 'YJLV', 'YJLY', 'VV', 'VLV', 'BBTRZ', 'RTTZ']): return '低压电力电缆'
        return '其他'

    from model_interpreter import interpret_model_short

    row = 2
    seq = 0
    overall_total = 0

    for cat in CATEGORIES:
        items = [a for a in aggregated if _categorize(a['model']) == cat]
        if not items:
            continue

        # 分类标题行
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        c = ws.cell(row=row, column=1, value=f'▸ {cat}')
        c.font = BOLD_FONT
        c.fill = FILL_CAT
        c.alignment = LEFT
        c.border = BORDER
        for ci in range(2, 9):
            ws.cell(row=row, column=ci).border = BORDER
        ws.row_dimensions[row].height = 24
        row += 1

        cat_total = 0
        for a in items:
            seq += 1
            _cell(ws, row, 1, seq, FILL_WHITE, CENTER)
            _cell(ws, row, 2, a['model'], FILL_WHITE, LEFT)

            # 来源行 → 超链接到 Sheet2
            src_rows = a.get('source_rows', [])
            src_str = ', '.join(str(s) for s in src_rows)
            ws.cell(row=row, column=6).value = src_str
            ws.cell(row=row, column=6).font = LINK_FONT
            ws.cell(row=row, column=6).alignment = LEFT
            # Make first source row clickable
            if src_rows:
                link_to_sheet2(ws, row=row, col=6, target_row=src_rows[0], display=src_str)

            _cell(ws, row, 3, 'm', FILL_WHITE, CENTER)
            _cell(ws, row, 4, _fmt_qty(a['qty']), FILL_WHITE, CENTER)

            # 型号解释（简版）
            short = interpret_model_short(a['model'])
            _cell(ws, row, 5, short, FILL_WHITE, LEFT)

            # 数量验算式
            expr = a.get('verification_expr', '')
            _cell(ws, row, 7, expr, FILL_WHITE, LEFT)
            _cell(ws, row, 8, a.get('note', ''), FILL_WHITE, LEFT)

            cat_total += a['qty']
            ws.row_dimensions[row].height = 20
            row += 1

        # 分类小计
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        c = ws.cell(row=row, column=1, value='  小计')
        c.font = BOLD_FONT; c.fill = FILL_ALT; c.alignment = LEFT; c.border = BORDER
        c = ws.cell(row=row, column=4, value=_fmt_qty(cat_total))
        c.font = BOLD_FONT; c.fill = FILL_ALT; c.alignment = CENTER; c.border = BORDER
        for ci in [5, 6, 7, 8]:
            ws.cell(row=row, column=ci).border = BORDER
        ws.row_dimensions[row].height = 22
        overall_total += cat_total
        row += 1

    # 合计
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    c = ws.cell(row=row, column=1, value='合  计')
    c.font = BOLD_FONT; c.fill = FILL_TOTAL; c.alignment = LEFT; c.border = BORDER
    c = ws.cell(row=row, column=4, value=_fmt_qty(overall_total))
    c.font = BOLD_FONT; c.fill = FILL_TOTAL; c.alignment = CENTER; c.border = BORDER
    for ci in [5, 6, 7, 8]:
        ws.cell(row=row, column=ci).border = BORDER
    ws.row_dimensions[row].height = 26

    ws.freeze_panes = 'A2'


# ════════════════════════════════════════════════════════════════
# Sheet4  型号解释
# ════════════════════════════════════════════════════════════════

def _build_sheet4(wb, aggregated):
    """非专业人员的通俗型号解释"""
    ws = wb.create_sheet('型号解释')

    headers = [
        ('规格型号', 44), ('一句话解释', 46), ('绝缘/护套材质', 24),
        ('阻燃耐火等级', 18), ('是否无卤', 10), ('适用场景', 24), ('注意事项', 32),
    ]
    _write_header_row(ws, headers, row=1)

    from model_interpreter import interpret_model_simple

    for i, a in enumerate(aggregated, start=2):
        info = interpret_model_simple(a['model'])
        _cell(ws, i, 1, a['model'], FILL_WHITE, LEFT)
        _cell(ws, i, 2, info.get('plain', ''), FILL_WHITE, LEFT)
        _cell(ws, i, 3, info.get('insulation', ''), FILL_WHITE, LEFT)
        _cell(ws, i, 4, info.get('fire_rating', ''), FILL_WHITE, CENTER)
        _cell(ws, i, 5, info.get('halogen_free', '—'), FILL_WHITE, CENTER)
        _cell(ws, i, 6, info.get('scenario', ''), FILL_WHITE, LEFT)
        _cell(ws, i, 7, info.get('caution', ''), FILL_WHITE, LEFT)
        ws.row_dimensions[i].height = 22

    ws.freeze_panes = 'A2'


# ════════════════════════════════════════════════════════════════
# Sheet5  待处理项
# ════════════════════════════════════════════════════════════════

def _build_sheet5(wb, issues):
    """需要人工介入的问题汇总"""
    ws = wb.create_sheet('待处理项')

    headers = [
        ('类型', 10), ('原始型号', 36), ('问题描述', 40),
        ('建议处理方式', 36), ('处理结果（人工填写）', 28),
    ]
    _write_header_row(ws, headers, row=1)

    for i, iss in enumerate(issues, start=2):
        fill = _issue_fill(iss['type'])
        _cell(ws, i, 1, iss['type'], fill, CENTER, font=BOLD_FONT)
        _cell(ws, i, 2, iss.get('raw_spec', ''), fill, LEFT)
        _cell(ws, i, 3, iss.get('description', ''), fill, LEFT)
        _cell(ws, i, 4, iss.get('suggestion', ''), fill, LEFT)
        # 结果列留空，浅黄色底色提示填写
        result_fill = PatternFill('solid', fgColor='FFFDE7')
        _cell(ws, i, 5, '', result_fill, LEFT, font=Font(name='Arial', size=10, color='999999', italic=True))
        ws.row_dimensions[i].height = 24

    ws.freeze_panes = 'A2'


def _issue_fill(typ):
    if '❌' in typ: return PatternFill('solid', fgColor='FFEBEE')
    if '⚠' in typ: return PatternFill('solid', fgColor='FFF3E0')
    if '❓' in typ: return PatternFill('solid', fgColor='FFFDE7')
    if '✏' in typ: return PatternFill('solid', fgColor='E8F5E9')
    return FILL_WHITE


# ════════════════════════════════════════════════════════════════
# Sheet6  操作日志
# ════════════════════════════════════════════════════════════════

def _build_sheet6(wb, log_entries):
    """审计追溯日志"""
    ws = wb.create_sheet('操作日志')

    headers = [
        ('时间戳', 18), ('处理阶段', 10), ('条目序号', 10),
        ('操作描述', 48), ('规则依据', 20),
    ]
    _write_header_row(ws, headers, row=1)

    now = datetime.now().strftime('%Y-%m-%d %H:%M')

    for i, ent in enumerate(log_entries, start=2):
        _cell(ws, i, 1, ent.get('timestamp', now), FILL_WHITE, CENTER)
        _cell(ws, i, 2, ent.get('phase', ''), FILL_WHITE, CENTER)
        _cell(ws, i, 3, ent.get('item_ref', ''), FILL_WHITE, CENTER)
        _cell(ws, i, 4, ent.get('operation', ''), FILL_WHITE, LEFT)
        _cell(ws, i, 5, ent.get('rule', ''), FILL_WHITE, LEFT)
        ws.row_dimensions[i].height = 20

    ws.freeze_panes = 'A2'


# ════════════════════════════════════════════════════════════════
# 辅助函数
# ════════════════════════════════════════════════════════════════

def _write_header_row(ws, headers, row=1):
    """写表头行"""
    for col, (text, width) in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=text)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = CENTER
        c.border = BORDER
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[row].height = 28


def _cell(ws, row, col, value, fill, alignment, font=None):
    """快速写单元格"""
    c = ws.cell(row=row, column=col, value=value)
    c.font = font or BODY_FONT
    c.fill = fill
    c.alignment = alignment
    c.border = BORDER
    return c


def _fmt_qty(val):
    """格式化数量：整数去小数，大数加千分位"""
    if val is None:
        return 0
    if isinstance(val, float) and val == int(val):
        return int(val)
    return val


def _set_col_widths(ws, widths: dict):
    """设置列宽"""
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def _set_auto_filter(ws, max_row, max_col):
    """设置自动筛选"""
    last_col = get_column_letter(max_col)
    ws.auto_filter.ref = f'A1:{last_col}{max_row}'


def link_to_sheet1(ws, *, row, col, target_row, display):
    """创建指向 Sheet1 的超链接"""
    _make_hyperlink(ws, row, col, f"#'原始清单'!A{target_row}", display)


def link_to_sheet2(ws, *, row, col, target_row, display):
    """创建指向 Sheet2 的超链接"""
    _make_hyperlink(ws, row, col, f"#'电缆提取+标准化'!A{target_row}", display)


def _make_hyperlink(ws, row, col, link_target, display):
    """内部超链接"""
    from openpyxl.worksheet.hyperlink import Hyperlink
    c = ws.cell(row=row, column=col)
    c.value = display
    c.hyperlink = link_target
    c.font = LINK_FONT
    c.alignment = LEFT


# ════════════════════════════════════════════════════════════════
# 向后兼容
# ════════════════════════════════════════════════════════════════

def to_excel(aggregated: dict, output_path: str, index_log: list = None,
             records: list = None) -> int:
    """
    旧版 to_excel — 简化输出（仅 Sheet3 风格）。
    """
    from aggregator import sort_key
    from model_interpreter import interpret_model_short

    wb = Workbook()
    ws = wb.active
    ws.title = '标准化采购清单'

    headers = [
        ('序号', 8), ('规格型号（标准化）', 48), ('单位', 6),
        ('数量', 12), ('型号解释（简版）', 40), ('备注', 20),
    ]
    _write_header_row(ws, headers, row=1)

    items = sorted(aggregated.items(), key=lambda x: sort_key(x[0]))
    for i, (model, qty) in enumerate(items, start=2):
        _cell(ws, i, 1, i - 1, FILL_WHITE, CENTER)
        _cell(ws, i, 2, model, FILL_WHITE, LEFT)
        _cell(ws, i, 3, 'm', FILL_WHITE, CENTER)
        _cell(ws, i, 4, _fmt_qty(qty), FILL_WHITE, CENTER)
        _cell(ws, i, 5, interpret_model_short(model), FILL_WHITE, LEFT)
        _cell(ws, i, 6, '', FILL_WHITE, LEFT)

    total = sum(aggregated.values())
    r = len(items) + 2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    _cell(ws, r, 1, '合  计', FILL_TOTAL, LEFT, font=BOLD_FONT)
    _cell(ws, r, 4, _fmt_qty(total), FILL_TOTAL, CENTER, font=BOLD_FONT)

    ws.freeze_panes = 'A2'
    wb.save(output_path)
    return len(items)
